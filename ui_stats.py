from datetime import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from PyQt6.QtCore import QDate, QLocale, Qt
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import (
    QCalendarWidget,
    QDateEdit,
    QDialog,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QVBoxLayout,
)

import database


class StatsDialog(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📊 Статистика и отчеты")
        self.resize(520, 400)

        self.init_ui()
        self.calculate_stats()

    def init_ui(self):
        main_layout = QVBoxLayout(self)

        # ---------------- 1. ВЫБОР ПЕРИОДА (КАЛЕНДАРЬ) ----------------
        period_group = QGroupBox("Выберите период")
        period_layout = QHBoxLayout(period_group)

        # Полный изолированный CSS для календаря, блокирующий перехват стилей QTableView из темы
        calendar_style = """
            QCalendarWidget QWidget {
                background-color: #ffffff !important;
                color: #1a1a1a !important;
            }
            QCalendarWidget QTableView {
                background-color: #ffffff !important;
                color: #1a1a1a !important;
                selection-background-color: #0067c0 !important;
                selection-color: #ffffff !important;
                gridline-color: #e5e5e5 !important;
            }
            QCalendarWidget QAbstractItemView:enabled {
                color: #1a1a1a !important;
                background-color: #ffffff !important;
                selection-background-color: #0067c0 !important;
                selection-color: #ffffff !important;
            }
            QCalendarWidget QAbstractItemView:disabled {
                color: #a0a0a0 !important;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #f3f3f3 !important;
            }
            QCalendarWidget QToolButton {
                color: #1a1a1a !important;
                background-color: #f3f3f3 !important;
            }
            QCalendarWidget QMenu {
                background-color: #ffffff !important;
                color: #1a1a1a !important;
            }
            QCalendarWidget QSpinBox {
                background-color: #ffffff !important;
                color: #1a1a1a !important;
            }
        """

        date_edit_style = """
            QDateEdit {
                min-width: 120px;
                padding: 5px 8px;
                font-size: 13px;
                font-weight: bold;
                color: #1a1a1a;
                background-color: #ffffff;
                border: 1px solid #d1d1d1;
                border-radius: 4px;
            }
            QDateEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 22px;
                border-left: 1px solid #d1d1d1;
            }
        """

        label_from = QLabel("С:")
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd.MM.yyyy")
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        self.date_from.setStyleSheet(date_edit_style)

        # Создаем явный виджет календаря с жестким стилем
        cal_from = QCalendarWidget(self)
        cal_from.setGridVisible(True)
        cal_from.setLocale(QLocale(QLocale.Language.Russian, QLocale.Country.Russia))
        cal_from.setStyleSheet(calendar_style)
        self.date_from.setCalendarWidget(cal_from)

        label_to = QLabel("по:")
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd.MM.yyyy")
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setStyleSheet(date_edit_style)

        cal_to = QCalendarWidget(self)
        cal_to.setGridVisible(True)
        cal_to.setLocale(QLocale(QLocale.Language.Russian, QLocale.Country.Russia))
        cal_to.setStyleSheet(calendar_style)
        self.date_to.setCalendarWidget(cal_to)

        btn_calc = QPushButton("🔄 Рассчитать")
        btn_calc.clicked.connect(self.calculate_stats)

        period_layout.addWidget(label_from)
        period_layout.addWidget(self.date_from)
        period_layout.addWidget(label_to)
        period_layout.addWidget(self.date_to)
        period_layout.addWidget(btn_calc)

        main_layout.addWidget(period_group)

        # ---------------- 2. РЕЗУЛЬТАТЫ СТАТИСТИКИ ----------------
        stats_group = QGroupBox("Итоги за период")
        stats_layout = QFormLayout(stats_group)

        self.lbl_gibdd_count = QLabel("0 шт.")
        self.lbl_gibdd_sum = QLabel("0.00 руб.")
        self.lbl_weapon_count = QLabel("0 шт.")
        self.lbl_weapon_sum = QLabel("0.00 руб.")

        self.lbl_total_count = QLabel("0 шт.")
        self.lbl_total_count.setStyleSheet("font-weight: bold; font-size: 14px;")

        self.lbl_total_sum = QLabel("0.00 руб.")
        self.lbl_total_sum.setStyleSheet("font-weight: bold; font-size: 15px; color: #0067c0;")

        stats_layout.addRow("Справок ГИБДД:", self.lbl_gibdd_count)
        stats_layout.addRow("Сумма ГИБДД:", self.lbl_gibdd_sum)
        stats_layout.addRow("------------------", QLabel(""))
        stats_layout.addRow("Справок на Оружие:", self.lbl_weapon_count)
        stats_layout.addRow("Сумма Оружие:", self.lbl_weapon_sum)
        stats_layout.addRow("==================", QLabel(""))
        stats_layout.addRow("Всего справок:", self.lbl_total_count)
        stats_layout.addRow("ИТОГО ПОЛУЧЕНО:", self.lbl_total_sum)

        main_layout.addWidget(stats_group)

        # ---------------- 3. КНОПКИ ЭКСПОРТА И ЗАКРЫТИЯ ----------------
        btn_export = QPushButton("📄 Выгрузить в Excel")
        btn_export.setStyleSheet("background-color: #28a745; color: white; font-weight: bold; padding: 8px;")
        btn_export.clicked.connect(self.export_to_excel)

        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(self.accept)

        btn_box = QHBoxLayout()
        btn_box.addWidget(btn_export)
        btn_box.addStretch()
        btn_box.addWidget(btn_close)
        main_layout.addLayout(btn_box)

    def calculate_stats(self):
        start_str = self.date_from.date().toString("yyyy-MM-dd")
        end_str = self.date_to.date().toString("yyyy-MM-dd")

        res = database.get_statistics_for_period(start_str, end_str)

        self.lbl_gibdd_count.setText(f"{res['gibdd_count']} шт.")
        self.lbl_gibdd_sum.setText(f"{res['gibdd_sum']:,.2f} руб.".replace(",", " "))

        self.lbl_weapon_count.setText(f"{res['weapon_count']} шт.")
        self.lbl_weapon_sum.setText(f"{res['weapon_sum']:,.2f} руб.".replace(",", " "))

        self.lbl_total_count.setText(f"{res['total_count']} шт.")
        self.lbl_total_sum.setText(f"{res['total_sum']:,.2f} руб.".replace(",", " "))

    def export_to_excel(self):
        start_date = self.date_from.date().toString("yyyy-MM-dd")
        end_date = self.date_to.date().toString("yyyy-MM-dd")

        if not hasattr(database, "get_spreadsheet_data_for_period"):
            QMessageBox.warning(
                self,
                "Внимание",
                "Функция выгрузки отчета не найдена в модуле database.",
            )
            return

        data = database.get_spreadsheet_data_for_period(start_date, end_date)

        if not data:
            QMessageBox.information(
                self, "Информация", "Нет данных за выбранный период."
            )
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Справки"

        headers = [
            "Номер выданной справки",
            "Дата выдачи",
            "До какого действительна",
            "Фамилия",
            "Имя",
            "Отчество",
            "Дата рождения",
            "Адрес",
            "Заключение (годен/не годен)",
            "Примечание",
        ]

        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

        column_widths = [15, 12, 15, 20, 15, 20, 12, 30, 15, 25]
        for i, width in enumerate(column_widths, 1):
            col_letter = (
                chr(64 + i)
                if i <= 26
                else f"{chr(64 + i // 26)}{chr(64 + i % 26)}"
            )
            ws.column_dimensions[col_letter].width = width

        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        start_fmt = self.date_from.date().toString("ddMMyyyy")
        end_fmt = self.date_to.date().toString("ddMMyyyy")
        default_filename = f"otchet_{start_fmt}_{end_fmt}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить отчет",
            default_filename,
            "Excel файлы (*.xlsx);;Все файлы (*.*)",
        )

        if file_path:
            try:
                wb.save(file_path)
                QMessageBox.information(
                    self, "Успех", f"Отчет успешно сохранен:\n{file_path}"
                )
            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось сохранить файл:\n{str(e)}"
                )